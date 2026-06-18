import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def read_csv(filepath: str) -> list[dict]:
    """
    Read a CSV or Excel (.xlsx, .xls) file containing prospects and return a list of standardized dictionaries.
    Handles decorative headers, empty rows, and maps columns like 'Founder(s)' or 'LinkedIn Page' automatically.
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"File not found at: {filepath}")
    
    # Load into DataFrame based on extension
    if filepath.lower().endswith(('.xlsx', '.xls')):
        df = pd.read_excel(filepath)
    else:
        df = pd.read_csv(filepath)
        
    # Standard keywords to find the true header row
    header_keywords = ["company", "name", "founder", "email", "linkedin", "title", "role"]
    
    # Let's see if we need to skip decorative top rows to find the actual header
    header_idx = -1
    for idx, row in df.iterrows():
        row_values = [str(val).lower() for val in row.values]
        matches = sum(1 for kw in header_keywords if any(kw in val for val in row_values))
        if matches >= 2:
            header_idx = idx
            break
            
    # Re-slice if we found a true header row lower down
    if header_idx != -1:
        new_header = [str(col).strip() for col in df.iloc[header_idx].values]
        df.columns = new_header
        df = df.iloc[header_idx + 1:]
        
    df = df.fillna("")
    
    # Map column headers to standard keys
    standard_map = {}
    for col in df.columns:
        col_lower = str(col).lower()
        if "name" in col_lower or "founder" in col_lower:
            standard_map[col] = "name"
        elif "email" in col_lower:
            standard_map[col] = "email"
        elif "company" in col_lower:
            standard_map[col] = "company"
        elif "title" in col_lower or "role" in col_lower:
            standard_map[col] = "title"
        elif "linkedin" in col_lower:
            standard_map[col] = "linkedin_url"
            
    df = df.rename(columns=standard_map)
    
    # Ensure all required standard fields exist in the output list
    required_cols = ["name", "email", "company", "title", "linkedin_url"]
    for col in required_cols:
        if col not in df.columns:
            df[col] = ""
            
    # Clean up string entries
    for col in required_cols:
        df[col] = df[col].astype(str).str.strip()
        
    # Remove rows where both name and company are completely empty
    df = df[~((df["name"] == "") & (df["company"] == ""))]
    
    return df.to_dict(orient="records")

def write_excel(prospects: list[dict], filepath: str):
    """
    Write a list of prospect dictionaries to an Excel file with styling.
    """
    if not prospects:
        return
    
    # Convert to DataFrame
    df = pd.DataFrame(prospects)
    
    # Rename columns to match user's custom requirements
    column_mapping = {
        "email": "Custom Mail",
        "email_subject": "Custom Subject",
        "email_body": "Custom Body"
    }
    df = df.rename(columns=column_mapping)
    
    # Ensure all required standard fields exist
    for col in ["Custom Mail", "Custom Subject", "Custom Body"]:
        if col not in df.columns:
            df[col] = ""
            
    # Filter to keep ONLY these three columns
    df = df[["Custom Mail", "Custom Subject", "Custom Body"]]
    
    # Create a new workbook and select active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Cold Email Campaign"
    
    # Colors & Fonts (Premium Palette: Navy Header with Mint accents)
    header_fill = PatternFill(start_color="1A2B4C", end_color="1A2B4C", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    data_font = Font(name="Segoe UI", size=10)
    success_fill = PatternFill(start_color="E2F8F0", end_color="E2F8F0", fill_type="solid")
    success_font = Font(name="Segoe UI", size=10, color="0D5C3A", bold=True)
    
    error_fill = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")
    error_font = Font(name="Segoe UI", size=10, color="9B1C1C", bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    # Write Headers
    ws.append(list(df.columns))
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = thin_border
    
    # Write Data
    for row_idx, row_data in enumerate(df.values, start=2):
        ws.append(list(row_data))
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            
            # Wrap text for long content (like email bodies or summaries)
            col_name = df.columns[col_idx - 1]
            if col_name in ["Custom Body", "email_body", "prospect_summary", "company_summary", "recent_news", "pain_points", "proofread_critique"]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="center")
            
            # Highlight status column
            if col_name == "status":
                val = str(cell.value).lower()
                if val in ["completed", "approved", "sent"]:
                    cell.fill = success_fill
                    cell.font = success_font
                elif val in ["failed", "rejected"]:
                    cell.fill = error_fill
                    cell.font = error_font
                    
    # Auto-adjust column width with padding
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        col_name = col[0].value
        
        # Don't make email_body and long summaries extremely wide
        if col_name in ["Custom Body", "Custom Subject", "email_body", "prospect_summary", "company_summary", "proofread_critique", "recent_news"]:
            ws.column_dimensions[col_letter].width = 45
            continue
            
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Set row height
    ws.row_dimensions[1].height = 25
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 20
        
    # Save the file
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
